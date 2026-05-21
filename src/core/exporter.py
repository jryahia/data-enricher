"""Export enriched data to CSV, JSON, or Excel."""
import csv
import json
from pathlib import Path
from typing import List, Dict, Any, Union


def export_csv(data: List[Dict[str, Any]], filepath: Union[str, Path]) -> str:
    """Export to CSV."""
    path = Path(filepath)
    if not data:
        raise ValueError("No data to export")
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    return str(path)


def export_json(data: List[Dict[str, Any]], filepath: Union[str, Path]) -> str:
    """Export to JSON."""
    path = Path(filepath)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    return str(path)


def export_xlsx(data: List[Dict[str, Any]], filepath: Union[str, Path]) -> str:
    """Export to Excel (.xlsx)."""
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")

    path = Path(filepath)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Enriched Data"

    if not data:
        raise ValueError("No data to export")

    # Write header
    headers = list(data[0].keys())
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    # Write data
    for row_idx, row in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            ws.cell(row=row_idx, column=col_idx, value=str(row.get(header, "")))

    wb.save(str(path))
    return str(path)


def export_data(
    data: List[Dict[str, Any]],
    filepath: Union[str, Path],
    fmt: str = "csv",
) -> str:
    """Export data in the specified format."""
    ext = Path(filepath).suffix.lower().lstrip(".")
    fmt = fmt.lower() if fmt else ext
    if fmt in ("csv",):
        return export_csv(data, filepath)
    elif fmt in ("json",):
        return export_json(data, filepath)
    elif fmt in ("xlsx", "xls", "excel"):
        return export_xlsx(data, filepath)
    else:
        # Auto-detect from extension
        if ext == "json":
            return export_json(data, filepath)
        elif ext in ("xlsx", "xls"):
            return export_xlsx(data, filepath)
        else:
            return export_csv(data, filepath)
